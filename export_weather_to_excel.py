#!/usr/bin/env python3
"""
Export Wetterbuch weather data to Excel file with monthly summaries.
Creates an Excel file with yearly and monthly weather statistics.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

# Weather types aligned with Wetterbuch.html (label, icon, color)
WEATHER_TYPES = [
    ('Sonne', '☀', '#f59e0b'),
    ('Heiter', '⛅', '#fbbf24'),
    ('Bewölkt', '☁', '#60a5fa'),
    ('Regen', '🌧', '#0369a1'),
    ('Gewitter', '⚡', '#1e3a8a'),
    ('Hagel', '🌨', '#d1d5db'),
    ('Schnee', '❄', '#7dd3fc'),
    ('Nebel', '🌫️', '#9ca3af'),
]


def hex_to_argb(hex_color):
    """Convert CSS hex color (#RRGGBB) to openpyxl ARGB format (FFRRGGBB)."""
    return f"FF{hex_color.lstrip('#').upper()}"


def set_weather_icon_cell(cell, weather_idx):
    """Write a weather icon into a cell using website-like coloring."""
    if weather_idx == 1:
        # Heiter: explicit two-tone symbol (yellow sun + gray cloud)
        sun = InlineFont(rFont='Segoe UI Symbol', sz=12, color=hex_to_argb('#fbbf24'))
        cloud = InlineFont(rFont='Segoe UI Symbol', sz=12, color=hex_to_argb('#94a3b8'))
        cell.value = CellRichText(TextBlock(sun, '☀'), TextBlock(cloud, '☁'))
        return

    _, icon, color = WEATHER_TYPES[weather_idx]
    cell.value = icon
    cell.font = Font(name='Segoe UI Emoji', size=12, color=hex_to_argb(color))

MONTHS_DE = [
    'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni',
    'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember'
]

MONTH_PREFIXES = {
    'JANUAR': 0, 'FEBRUAR': 1, 'MAERZ': 2, 'APRIL': 3,
    'MAI': 4, 'JUNI': 5, 'JULI': 6, 'AUGUST': 7,
    'SEPTEMBER': 8, 'OKTOBER': 9, 'NOVEMBER': 10, 'DEZEMBER': 11,
}

def detect_month_from_filename(filename):
    """Extract month index from filename."""
    up = filename.upper()
    for prefix, idx in MONTH_PREFIXES.items():
        if up.startswith(prefix):
            return idx
    return -1

def parse_seq_file(buffer):
    """Parse SEQ file containing weather data."""
    try:
        text = buffer.decode('latin1', errors='ignore')
    except:
        return None
    
    # C64 PRINT# separates values with CR (0x0D)
    parts = text.split('\r')
    parts = [s.strip() for s in parts if s.strip()]
    
    if len(parts) < 16:
        return None
    
    try:
        summary_raw = [int(p) for p in parts[-16:]]
        day_raw = [int(p) for p in parts[:-16]]
    except ValueError:
        return None
    
    days = []
    for i in range(0, len(day_raw) - 4, 5):
        days.append({
            'minTemp': day_raw[i],
            'maxTemp': day_raw[i + 1],
            'weather': day_raw[i + 2],
            'rain': day_raw[i + 3],
            'snow': day_raw[i + 4],
        })
    
    # If summary is all zeros (incomplete save), recompute from daily data
    if all(v == 0 for v in summary_raw[:8]) and days:
        summary_raw = compute_summary(days)
    
    return {
        'days': days,
        'summary': {
            'maxTemp': summary_raw[0],
            'minTemp': summary_raw[2],
            'maxRain': summary_raw[4],
            'maxSnow': summary_raw[6],
            'weatherCounts': summary_raw[8:16],
        }
    }

def compute_summary(days):
    """Compute monthly summary from daily data."""
    max_temp, min_temp = -999, 999
    max_rain, max_snow = 0, 0
    weather_counts = [0] * 8
    
    for day in days:
        max_temp = max(max_temp, day['maxTemp'])
        min_temp = min(min_temp, day['minTemp'])
        max_rain = max(max_rain, day['rain'])
        max_snow = max(max_snow, day['snow'])
        
        # Count weather types from bitmask
        for bit in range(8):
            if day['weather'] & (1 << bit):
                weather_counts[bit] += 1
    
    return [max_temp, 0, min_temp, 0, max_rain, 0, max_snow, 0] + weather_counts

def get_weather_mask(summary):
    """Get bitmask of which weather types occurred in the month."""
    weather_counts = summary.get('weatherCounts', [0] * 8)
    mask = 0
    for i, count in enumerate(weather_counts):
        if count > 0:
            mask |= (1 << i)
    return mask

def export_to_excel(output_file='Wetterdaten.xlsx'):
    """Export all weather data to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Wetterbuch_Monat'
    ws_day = wb.create_sheet('Wetterbuch_Tag')
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # Jahr
    ws.column_dimensions['B'].width = 12  # Monat
    for col in range(3, 11):  # Weather icons
        ws.column_dimensions[get_column_letter(col)].width = 5
    ws.column_dimensions['K'].width = 10  # Min Temp
    ws.column_dimensions['L'].width = 10  # Max Temp
    ws.column_dimensions['M'].width = 12  # Regen(mm)
    ws.column_dimensions['N'].width = 12  # Schnee(cm)

    # Set column widths for daily sheet
    ws_day.column_dimensions['A'].width = 8   # Jahr
    ws_day.column_dimensions['B'].width = 12  # Monat
    ws_day.column_dimensions['C'].width = 6   # Tag
    for col in range(4, 12):  # Weather icons
        ws_day.column_dimensions[get_column_letter(col)].width = 5
    ws_day.column_dimensions['L'].width = 10  # Min Temp
    ws_day.column_dimensions['M'].width = 10  # Max Temp
    ws_day.column_dimensions['N'].width = 12  # Regen(mm)
    ws_day.column_dimensions['O'].width = 12  # Schnee(cm)
    
    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    headers = ['Jahr', 'Monat'] + [name for name, _, _ in WEATHER_TYPES] + ['Min Temp', 'Max Temp', 'Regen(mm)', 'Schnee(cm)']
    ws.append(headers)

    day_headers = ['Jahr', 'Monat', 'Tag'] + [name for name, _, _ in WEATHER_TYPES] + ['Min Temp', 'Max Temp', 'Regen(mm)', 'Schnee(cm)']
    ws_day.append(day_headers)
    
    # Format header row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    for col in range(1, len(day_headers) + 1):
        cell = ws_day.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    ws_day.freeze_panes = 'A2'
    
    # Collect data for all years and months
    data_rows = []
    day_rows = []
    
    for year in range(1985, 1999):
        year_dir = Path(f'{year}')
        if not year_dir.exists():
            continue
        
        # Collect month files
        month_files = {}
        for seq_file in year_dir.glob('*.SEQ'):
            month_idx = detect_month_from_filename(seq_file.name)
            if month_idx >= 0:
                month_files[month_idx] = seq_file
        
        # Process each month
        for month_idx in range(12):
            if month_idx not in month_files:
                continue
            
            seq_path = month_files[month_idx]
            
            # Read and parse the file
            try:
                with open(seq_path, 'rb') as f:
                    buffer = f.read()
                parsed = parse_seq_file(buffer)
                if not parsed:
                    continue
                
                summary = parsed['summary']
                weather_mask = get_weather_mask(summary)
                
                # Build row
                row = [year, MONTHS_DE[month_idx]]
                
                # Weather icons (columns 3-10)
                for bit in range(8):
                    if weather_mask & (1 << bit):
                        row.append(WEATHER_TYPES[bit][1])
                    else:
                        row.append('')
                
                # Temperature and precipitation
                row.extend([
                    summary['minTemp'],
                    summary['maxTemp'],
                    summary['maxRain'],
                    summary['maxSnow'],
                ])
                
                data_rows.append(row)

                # Build daily rows (second tab)
                for day_idx, day in enumerate(parsed['days'], start=1):
                    day_row = [year, MONTHS_DE[month_idx], day_idx]
                    day_mask = day['weather']
                    for bit in range(8):
                        if day_mask & (1 << bit):
                            day_row.append(WEATHER_TYPES[bit][1])
                        else:
                            day_row.append('')
                    day_row.extend([
                        day['minTemp'],
                        day['maxTemp'],
                        day['rain'],
                        day['snow'],
                    ])
                    day_rows.append(day_row)
                
            except Exception as e:
                print(f'Error processing {seq_path}: {e}')
                continue
    
    # Write data rows
    for row_data in data_rows:
        ws.append(row_data)

    for row_data in day_rows:
        ws_day.append(row_data)
    
    # Format data rows
    for row_idx in range(2, len(data_rows) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = center_align
            
            # Format numbers
            if col_idx in [1, 11, 12, 13, 14]:  # Year and numeric columns
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply website-style weather icon colors in columns C-J
            if 3 <= col_idx <= 10 and cell.value:
                weather_idx = col_idx - 3
                set_weather_icon_cell(cell, weather_idx)

    # Format daily rows
    for row_idx in range(2, len(day_rows) + 2):
        for col_idx in range(1, len(day_headers) + 1):
            cell = ws_day.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = center_align

            # Year, day, and numeric columns
            if col_idx in [1, 3, 12, 13, 14, 15]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply website-style weather icon colors in columns D-K
            if 4 <= col_idx <= 11 and cell.value:
                weather_idx = col_idx - 4
                set_weather_icon_cell(cell, weather_idx)
    
    # Save the workbook
    wb.save(output_file)
    print(f'✅ Excel file created: {output_file}')
    print(f'   Monthly rows: {len(data_rows)} months from 14 years')
    print(f'   Daily rows: {len(day_rows)} days')

if __name__ == '__main__':
    export_to_excel('Wetterdaten.xlsx')
